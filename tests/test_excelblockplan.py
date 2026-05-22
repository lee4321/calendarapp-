"""Tests for the excelblockplan Excel workbook generator."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from config.config import create_calendar_config
from visualizers.excelblockplan import generate_excel_blockplan
from visualizers.excelheader import (
    CONTINUATION_COL,
    FIRST_DATE_COL,
    FIXED_COLUMNS,
    LABEL_COL_END,
)


# ── Stubs ─────────────────────────────────────────────────────────────────────


class _BaseDB:
    """Minimal DB stub — overridable hooks for events/holidays."""

    EVENTS: list[dict] = []
    HOLIDAYS: dict[str, list[dict]] = {}
    SPECIAL: dict[str, list[dict]] = {}
    NONWORK_KEYS: set[str] = set()

    def get_palette(self, name):  # noqa: D401
        return None

    def get_holidays_for_date(self, daykey, country=None):
        return list(self.HOLIDAYS.get(daykey, []))

    def get_special_days_for_date(self, daykey):
        return list(self.SPECIAL.get(daykey, []))

    def is_nonworkday(self, daykey, country=None):
        return daykey in self.NONWORK_KEYS

    def is_government_nonworkday(self, daykey, country=None):
        return daykey in {
            k for k, v in self.HOLIDAYS.items() if any(h.get("nonworkday") for h in v)
        }

    def get_all_events_in_range(self, start, end):
        return list(self.EVENTS)


def _event(
    *,
    eid: int,
    name: str,
    start: str,
    end: str | None = None,
    milestone: bool = False,
    priority: int = 0,
    icon: str = "",
    color: str = "",
    resource_group: str = "",
    wbs: str = "",
    rollup: bool = False,
    percent_complete: float = 0.0,
    status: str = "active",
    notes: str = "",
) -> dict:
    return {
        "ID": eid,
        "Status": status,
        "Priority": priority,
        "WBS": wbs,
        "Rollup": rollup,
        "Milestone": milestone,
        "Percent_Complete": percent_complete,
        "Task_Name": name,
        "Effort": 0,
        "Duration": 0,
        "Start_Date": start,
        "Finish_Date": end or start,
        "Earliest_Start_Date": "",
        "Latest_Start_Date": "",
        "Earliest_End_Date": "",
        "Latest_End_Date": "",
        "Predecessors": "",
        "Resource_Names": "",
        "Resource_Group": resource_group,
        "Notes": notes,
        "Icon": icon,
        "Color": color,
        "Tags": "",
        "Datekey": start,
        "Start": start,
        "End": end or start,
        "nonworkday": False,
    }


def _cfg(out_path: Path):
    c = create_calendar_config()
    c.outputfile = str(out_path)
    # Mon Jan 5 – Fri Jan 23 2026 (3 workweeks)
    c.userstart = "20260105"
    c.userend = "20260123"
    c.adjustedstart = "20260105"
    c.adjustedend = "20260123"
    c.weekend_style = 0
    c.country = None
    # The default config.rollups is True ("show only rollups") which would
    # filter out every test event. Tests want the full set.
    c.rollups = False
    c.excelheader_font_name = "Calibri"
    c.excelheader_font_size = 9
    c.excelheader_top_time_bands = [{"label": "Month", "unit": "month"}]
    return c


# ── Tests ─────────────────────────────────────────────────────────────────────


def test_excelblockplan_creates_file_with_planner_sheet(tmp_path):
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)
    generate_excel_blockplan(cfg, _BaseDB(), out)
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert "Planner" in wb.sheetnames


def test_excelblockplan_column_header_row_has_all_23_labels(tmp_path):
    """Row N+1 must populate columns A-W with the events-table field names."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)
    generate_excel_blockplan(cfg, _BaseDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    header_row = 2  # 1 timeband row + 1 column-header row
    expected = [name for name, _w in FIXED_COLUMNS]
    headers = [
        ws.cell(row=header_row, column=c).value for c in range(1, LABEL_COL_END + 1)
    ]
    assert headers == expected
    # X column reserved for the continuation icon — header cell is empty
    assert ws.cell(row=header_row, column=CONTINUATION_COL).value in ("", None)
    # First date column starts at Y (= FIRST_DATE_COL)
    assert get_column_letter(FIRST_DATE_COL) == "Y"


def test_excelblockplan_events_each_on_own_row_ordered_by_start(tmp_path):
    """Each event/duration occupies its own row; rows sorted by start date."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)

    class _DB(_BaseDB):
        EVENTS = [
            _event(eid=2, name="Zebra", start="20260120"),
            _event(eid=1, name="Apple", start="20260106", milestone=True),
            _event(eid=3, name="Mango", start="20260112", end="20260115"),
        ]

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    data_start = 3  # 1 timeband + 1 header
    name_col = 8  # column H
    names = [ws.cell(row=data_start + i, column=name_col).value for i in range(3)]
    assert names == ["Apple", "Mango", "Zebra"]
    # Each event got its own row — no merged data cells in this range.
    for row in range(data_start, data_start + 3):
        for col in range(1, LABEL_COL_END + 1):
            cell = ws.cell(row=row, column=col)
            assert not isinstance(cell, openpyxl.cell.cell.MergedCell), (
                f"cell {col},{row} is a merged cell — data rows must be unmerged"
            )


def test_excelblockplan_event_icon_placed_in_start_date_column(tmp_path):
    """Single-day events must place a glyph in the day-column for their start date."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)

    class _DB(_BaseDB):
        EVENTS = [
            _event(eid=1, name="Release", start="20260112", milestone=True, icon="★"),
        ]

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    # Mon Jan 12 is the 6th visible weekday (Jan 5-9 = 5 days, Jan 12 = 6th)
    event_col = FIRST_DATE_COL + 5
    data_row = 3
    cell = ws.cell(row=data_row, column=event_col)
    assert cell.value == "★"
    # Neighbouring day should be empty (no other events on that row)
    neighbour = ws.cell(row=data_row, column=event_col + 1)
    assert neighbour.value in (None, "")


def test_excelblockplan_duration_fills_day_columns_between_start_and_end(tmp_path):
    """Durations fill every visible day column between start and end."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)

    class _DB(_BaseDB):
        EVENTS = [
            _event(
                eid=1,
                name="Sprint 1",
                start="20260112",
                end="20260116",
                color="#4472C4",
            ),
        ]

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    data_row = 3
    # Jan 12 = visible day 5 (0-indexed); Jan 16 = visible day 9.
    for offset in range(5, 10):
        cell = ws.cell(row=data_row, column=FIRST_DATE_COL + offset)
        assert cell.fill.fill_type == "solid", (
            f"day offset {offset} should have a solid fill from the duration"
        )
    # Day immediately before start should NOT be filled.
    pre = ws.cell(row=data_row, column=FIRST_DATE_COL + 4)
    assert pre.fill.fill_type is None
    # Day immediately after end should NOT be filled.
    post = ws.cell(row=data_row, column=FIRST_DATE_COL + 10)
    assert post.fill.fill_type is None


def test_excelblockplan_holiday_overlays_duration_with_pattern(tmp_path):
    """Federal holidays must show through duration colour via a pattern fill."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)
    cfg.excelheader_federal_holiday_fill_color = "#FF0000"

    class _DB(_BaseDB):
        EVENTS = [
            _event(eid=1, name="Long task", start="20260112", end="20260123", color="#4472C4"),
        ]
        HOLIDAYS = {
            "20260119": [
                {"displayname": "MLK Day", "icon": "us", "nonworkday": 1, "country": "US"}
            ]
        }
        NONWORK_KEYS = {"20260119"}

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    # Mon Jan 19 → visible weekday offset 10
    holiday_col = FIRST_DATE_COL + 10
    data_row = 3
    cell = ws.cell(row=data_row, column=holiday_col)
    # Pattern fill that mixes overlay (federal red) with base (duration blue)
    assert cell.fill.fill_type == "lightUp", (
        f"expected lightUp pattern overlay on holiday inside duration, got {cell.fill.fill_type!r}"
    )


def test_excelblockplan_continuation_marker_for_duration_running_past_range(tmp_path):
    """When a duration extends beyond the visible range, column X carries a marker."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)

    class _DB(_BaseDB):
        EVENTS = [
            # Starts before, ends after — both arrows expected
            _event(eid=1, name="Forever", start="20251201", end="20260601"),
        ]

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    data_row = 3
    marker = ws.cell(row=data_row, column=CONTINUATION_COL).value
    assert marker in ("◀▶", "◀", "▶"), f"expected a continuation glyph, got {marker!r}"
    assert "◀" in marker and "▶" in marker


def test_excelblockplan_freeze_panes_at_first_date_column(tmp_path):
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)
    generate_excel_blockplan(cfg, _BaseDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws.freeze_panes == f"{get_column_letter(FIRST_DATE_COL)}2"


def test_excelblockplan_default_output_filename(tmp_path, monkeypatch):
    """The CLI default output filename must be ExcelBlockplan.xlsx under output/."""
    import ecalendar
    from pathlib import Path as _P

    # The CLI looks up calendar.db relative to cwd; point it at the repo copy
    # so the run does not error out on a missing database.
    repo_root = _P(__file__).resolve().parent.parent
    monkeypatch.chdir(tmp_path)
    db_path = repo_root / "calendar.db"
    rc = ecalendar.run(
        [
            "ecalendar.py", "excelblockplan",
            "20260105", "20260109",
            "--database", str(db_path),
            "--quiet",
        ]
    )
    assert rc == 0
    assert (tmp_path / "output" / "ExcelBlockplan.xlsx").exists()


def test_excelblockplan_content_filters_exclude_events(tmp_path):
    """--noevents flag should drop single-day events from the sheet."""
    out = tmp_path / "bp.xlsx"
    cfg = _cfg(out)
    cfg.includeevents = False
    cfg.includedurations = True

    class _DB(_BaseDB):
        EVENTS = [
            _event(eid=1, name="Solo event", start="20260108", milestone=True),
            _event(eid=2, name="Duration", start="20260112", end="20260116"),
        ]

    generate_excel_blockplan(cfg, _DB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    data_row = 3
    # First data row should be the duration (only surviving record)
    assert ws.cell(row=data_row, column=8).value == "Duration"
    # No second data row
    assert ws.cell(row=data_row + 1, column=8).value in (None, "")
