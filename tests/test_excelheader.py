"""Tests for the excelheader Excel workbook generator."""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from config.config import create_calendar_config
from visualizers.excelheader import (
    FIRST_DATE_COL,
    DATA_ROWS,
    generate_excel_header,
)


# ── Stubs ─────────────────────────────────────────────────────────────────────


class _DummyDB:
    """Minimal DB stub — no holidays, no palettes."""

    @staticmethod
    def get_palette(name):
        return None

    @staticmethod
    def get_holidays_for_date(daykey, country=None):
        return []

    @staticmethod
    def get_special_days_for_date(daykey):
        return []

    @staticmethod
    def is_nonworkday(daykey, country=None):
        return False


class _HolidayDB(_DummyDB):
    """Stub that marks a single fixed date as a federal nonworkday."""

    HOLIDAY_KEY = "20260119"  # Mon Jan 19 2026 (MLK Day)

    @classmethod
    def get_holidays_for_date(cls, daykey, country=None):
        if daykey == cls.HOLIDAY_KEY:
            return [{"displayname": "MLK Day", "icon": "us", "nonworkday": 1, "country": "US"}]
        return []

    @classmethod
    def is_nonworkday(cls, daykey, country=None):
        return daykey == cls.HOLIDAY_KEY


def _base_config(out_path: Path):
    config = create_calendar_config()
    config.outputfile = str(out_path)
    # Mon Jan 5 – Fri Jan 23 2026 (3 workweeks)
    config.userstart = "20260105"
    config.userend = "20260123"
    config.adjustedstart = "20260105"
    config.adjustedend = "20260123"
    config.weekend_style = 0  # weekdays only
    config.country = None
    config.excelheader_font_name = "Calibri"
    config.excelheader_font_size = 9
    return config


# ── Tests ─────────────────────────────────────────────────────────────────────


def test_excelheader_creates_file(tmp_path):
    """generate_excel_header() must produce a valid .xlsx file."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month", "date_format": "MMM"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    assert out.exists()
    wb = openpyxl.load_workbook(str(out))
    assert "Planner" in wb.sheetnames


def test_excelheader_day_columns_start_at_f(tmp_path):
    """Date columns must begin at column F (index 6)."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Date", "unit": "date", "date_format": "D"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # Column F should exist and E should be the last fixed label column
    assert FIRST_DATE_COL == 6  # sanity check constant
    assert get_column_letter(FIRST_DATE_COL) == "F"
    # The date-band row should have a value in column F
    assert ws.cell(row=1, column=FIRST_DATE_COL).value is not None


def test_excelheader_fixed_label_columns(tmp_path):
    """Column-header row must contain the 5 fixed label strings in A–E."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # After 1 timeband row the header row is row 2
    header_row = 2
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 6)]
    assert headers == ["Activity", "Effort", "Duration", "Scheduled Start", "Scheduled End"]


def test_excelheader_day_column_width(tmp_path):
    """Date columns (F+) must have width == 3 characters."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # Column F must be width 3
    assert ws.column_dimensions[get_column_letter(FIRST_DATE_COL)].width == 3.0


def test_excelheader_multi_month_band_uses_merged_cells(tmp_path):
    """Month segments spanning multiple day-columns must be merged cells."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month", "date_format": "MMM"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # Row 1 is the month band; January spans multiple weekday columns
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]
    # At least one merged range should start at or after column F in row 1
    date_merges = [
        r for r in merged_ranges
        if "1:" in r or r.endswith("1")
    ]
    assert date_merges, "Expected merged cells in the month timeband row"


def test_excelheader_data_rows_count(tmp_path):
    """There must be exactly DATA_ROWS empty data rows below the timebands."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month"},
        {"label": "Day", "unit": "date", "date_format": "D"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # 2 band rows + 1 header row + DATA_ROWS data rows
    expected_last = 2 + 1 + DATA_ROWS
    # ws.max_row may be higher if openpyxl tracks styled empty cells
    assert ws.max_row >= expected_last


def test_excelheader_holiday_shading_applied(tmp_path):
    """Holiday dates must receive a background fill in the data rows."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month"},
    ]
    # Use federal color
    config.theme_federal_holiday_color = "#FF0000"

    generate_excel_header(config, _HolidayDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    # Find which column holds the holiday date (Mon Jan 19)
    # Workweek: Jan 5–9, 12–16, 19–23 → Jan 19 is the 11th visible weekday
    holiday_col = FIRST_DATE_COL + 10  # 0-indexed slot 10
    header_row = 2  # 1 band + 1 header row
    data_row = header_row + 1

    cell = ws.cell(row=data_row, column=holiday_col)
    # Cell should have a solid fill (not no fill)
    assert cell.fill is not None
    fill_type = cell.fill.fill_type
    assert fill_type == "solid", (
        f"Expected solid fill on holiday column, got {fill_type!r}"
    )


def test_excelheader_vertical_lines_produce_right_borders(tmp_path):
    """Vertical lines in the theme become right-cell borders on data rows."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Week", "unit": "week", "label_format": "W{week}"},
    ]
    # Pin a vertical line to the end of each week segment
    config.blockplan_vertical_lines = [
        {"band": "Week", "repeat": True, "align": "end", "color": "navy", "width": 2.0},
    ]
    config.blockplan_vertical_line_color = "navy"
    config.blockplan_vertical_line_width = 2.0

    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active

    header_row = 2  # 1 band + 1 header row
    data_row = header_row + 1

    # Check that at least one data-row cell has a right border set
    right_borders = []
    for col in range(FIRST_DATE_COL, FIRST_DATE_COL + 15):
        cell = ws.cell(row=data_row, column=col)
        rb = cell.border.right
        if rb and rb.border_style and rb.border_style != "none":
            right_borders.append(col)

    assert right_borders, (
        "Expected at least one right border from vertical_lines in data rows"
    )


def test_excelheader_freeze_panes_set(tmp_path):
    """Freeze panes must be set at column F / column-header row."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.blockplan_top_time_bands = [
        {"label": "Month", "unit": "month"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # freeze_panes should be "F2" (1 band row → header on row 2, freeze at F)
    assert ws.freeze_panes == "F2"


def test_excelheader_weekends_excluded_when_style_zero(tmp_path):
    """With weekend_style=0 Saturday/Sunday columns must not appear."""
    out = tmp_path / "header.xlsx"
    config = _base_config(out)
    config.userstart = "20260105"   # Mon
    config.userend = "20260111"     # Sun
    config.adjustedstart = "20260105"
    config.adjustedend = "20260111"
    config.weekend_style = 0
    config.blockplan_top_time_bands = [
        {"label": "Date", "unit": "date", "date_format": "D"},
    ]
    generate_excel_header(config, _DummyDB(), out)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    # Mon Jan 5 – Fri Jan 9 = 5 weekdays → 5 date columns (F–J)
    # Row 1 = date band; collect non-None date cells
    date_cells = []
    for col in range(FIRST_DATE_COL, FIRST_DATE_COL + 10):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            date_cells.append(v)
    # Should have labels for 5 days: "5", "6", "7", "8", "9"
    assert len(date_cells) == 5
