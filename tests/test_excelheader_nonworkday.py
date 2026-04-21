"""Tests for excelheader non-workday fills using classify_day."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from config.config import create_calendar_config
from visualizers.excelheader import FIRST_DATE_COL, generate_excel_header


class _DummyDB:
    @staticmethod
    def get_palette(name):
        return None

    @staticmethod
    def get_holidays_for_date(daykey, country=None):
        return []

    @staticmethod
    def get_special_days_for_date(daykey):
        return []

    @staticmethod
    def is_nonworkday(daykey, country=None):
        return False

    @staticmethod
    def is_government_nonworkday(daykey, country=None):
        return False

    @staticmethod
    def get_all_events_in_range(start, end):
        return []


def _cfg(out_path: Path):
    c = create_calendar_config()
    c.outputfile = str(out_path)
    c.userstart = "20260302"   # Mon
    c.userend = "20260315"     # Sun — full 2 weeks including weekends
    c.adjustedstart = "20260302"
    c.adjustedend = "20260315"
    c.weekend_style = 1        # show all 7 weekdays
    c.weekend_days = [5, 6]    # Sat, Sun
    c.country = None
    c.excelheader_font_name = "Calibri"
    c.excelheader_font_size = 9
    return c


def test_weekend_fill_applied_when_configured(tmp_path):
    out = tmp_path / "eh.xlsx"
    cfg = _cfg(out)
    cfg.excelheader_weekend_fill_color = "#DDDDDD"
    cfg.excelheader_top_time_bands = [{"label": "Month", "unit": "month"}]
    generate_excel_header(cfg, _DummyDB(), out)

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    data_row = 3   # 1 band + 1 header + data_row start

    # Sat = 2026-03-07 → offset 5 (Mon=0..Sun=6) from Mar 2
    sat_col = FIRST_DATE_COL + 5
    cell = ws.cell(row=data_row, column=sat_col)
    assert cell.fill.fill_type == "solid"
    # ARGB is FF + DDDDDD
    assert cell.fill.start_color.rgb.upper() == "FFDDDDDD"


def test_weekend_fill_absent_without_config(tmp_path):
    out = tmp_path / "eh2.xlsx"
    cfg = _cfg(out)
    # No excelheader_weekend_fill_color set
    cfg.excelheader_top_time_bands = [{"label": "Month", "unit": "month"}]
    generate_excel_header(cfg, _DummyDB(), out)

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    data_row = 3
    sat_col = FIRST_DATE_COL + 5
    cell = ws.cell(row=data_row, column=sat_col)
    # No nonworkday config → no fill
    assert cell.fill.fill_type is None
