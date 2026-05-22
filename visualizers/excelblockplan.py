"""Excel workbook generator — blockplan-style data sheet.

Builds on the same A-W / Y+ skeleton produced by :mod:`visualizers.excelheader`
and fills in the data rows with one record per event and duration sourced from
the events table.

Layout reminder (see :mod:`visualizers.excelheader` for the full spec)
----------------------------------------------------------------------
Rows 1..N   : timeband rows — heading label in column W, segment values from Y
Row  N+1    : column-header row with the 23 events-table column names in A-W
Rows N+2..  : one row per event/duration after filtering, ordered by start_date

For each data row:
    - Columns A-W hold the corresponding events-table field values.
    - Column X holds a continuation marker when a duration extends past the
      visible range (left ◀ / right ▶ / both ◀▶ glyph).
    - Single-day events place the resolved icon name in the start-date day
      column (Y+).  The icon cell font colour and fill come from
      ``style_rules`` evaluated for the event.
    - Multi-day durations fill every day column between start and end with the
      style-resolved colour.
    - After all data is drawn, holiday/special-day shading is applied to the
      Y+ columns.  Cells that already carry data are decorated with a
      ``lightUp`` pattern that visibly combines the holiday colour and the
      underlying data colour so both remain visible.

The CLI mirrors ``blockplan`` (content filters, theme, weekends, country, etc.).
"""

from __future__ import annotations

from bisect import bisect_left
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING, Any

import arrow
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from shared.data_models import Event
from shared.rule_engine import DayContext, StyleEngine
from visualizers.base import filter_events
from visualizers.excelheader import (
    CONTINUATION_COL,
    FIRST_DATE_COL,
    FIXED_COLUMNS,
    LABEL_COL_END,
    _apply_fill,
    _apply_overlay_fill,
    _apply_right_border,
    _font_color_argb,
    _prepare_sheet,
    _to_argb,
    compute_visible_days,
)

if TYPE_CHECKING:
    from config.config import CalendarConfig
    from shared.db_access import CalendarDB


# Map FIXED_COLUMNS field-name → events-table dict key (as returned by
# ``CalendarDB.get_all_events_in_range``).  Keys missing from a row yield "".
_EVENT_FIELD_MAP: dict[str, str] = {
    "id": "ID",
    "status": "Status",
    "priority": "Priority",
    "wbs": "WBS",
    "rollup": "Rollup",
    "milestone": "Milestone",
    "percent_complete": "Percent_Complete",
    "name": "Task_Name",
    "effort": "Effort",
    "duration": "Duration",
    "start_date": "Start_Date",
    "end_date": "Finish_Date",
    "earliest_start_date": "Earliest_Start_Date",
    "latest_start_date": "Latest_Start_Date",
    "earliest_end_date": "Earliest_End_Date",
    "latest_end_date": "Latest_End_Date",
    "predecessors": "Predecessors",
    "resource_names": "Resource_Names",
    "resource_group": "Resource_Group",
    "notes": "Notes",
    "icon": "Icon",
    "color": "Color",
    "tags": "Tags",
}


def _format_cell_value(field_name: str, raw: Any) -> Any:
    """Coerce a raw events-table value into something readable in Excel.

    Numerics pass through as numbers; bools render as 1/0 so Excel formula
    bars stay friendly; ``None`` → ``""`` so cells appear blank rather than
    "None".  Date strings stay as ``YYYYMMDD`` text on purpose — adding Excel
    date-type coercion is out of scope and would prevent custom formatting on
    rows that mix populated and blank dates.
    """
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return 1 if raw else 0
    if field_name == "percent_complete" and isinstance(raw, (int, float)):
        return float(raw)
    return raw


def _event_day_context(event: Event) -> DayContext:
    """Build a DayContext for *event* using only event-known state.

    Day-classification keys (federal_holiday, weekend, etc.) are left at their
    defaults — style rules that key off non-workday context will still
    evaluate, but typically excelblockplan rules drive off event fields
    (resource_group, milestone, priority, …) so this is sufficient.
    """
    return DayContext(date=event.start or "")


def _resolve_event_style(
    engine: StyleEngine | None, event: Event
) -> tuple[str | None, str | None]:
    """Return ``(fill_color, icon_color)`` from style_rules for *event*.

    Falls back to the event's own ``color`` field when no rule supplies a fill.
    """
    fill_color: str | None = event.color or None
    icon_color: str | None = None
    if engine is not None:
        sr = engine.evaluate_event(event, ctx=_event_day_context(event))
        if sr.fill_color and not isinstance(sr.fill_color, list):
            fill_color = str(sr.fill_color)
        elif isinstance(sr.fill_color, list) and sr.fill_color:
            fill_color = str(sr.fill_color[0])
        if sr.icon_color:
            icon_color = str(sr.icon_color)
    if icon_color is None:
        icon_color = fill_color
    return fill_color, icon_color


def _resolve_event_icon(engine: StyleEngine | None, event: Event) -> str:
    """Return the icon glyph/name to draw for *event*.

    Priority: style_rule ``icon`` → events-table ``icon`` → ``"●"``.
    The value is rendered into the Excel cell verbatim — Unicode characters
    work naturally; named icons would need a font that ships with the user's
    Excel install, so we just write the string.
    """
    if engine is not None:
        sr = engine.evaluate_event(event, ctx=_event_day_context(event))
        if sr.icon:
            return str(sr.icon)
    if event.icon:
        return str(event.icon)
    return "●"


def _column_for_day(visible_days: list[date], d: date) -> int | None:
    """Return the 1-based Excel column for *d*, or ``None`` if not visible."""
    idx = bisect_left(visible_days, d)
    if 0 <= idx < len(visible_days) and visible_days[idx] == d:
        return FIRST_DATE_COL + idx
    return None


def _continuation_glyph(
    *, continues_left: bool, continues_right: bool
) -> str:
    if continues_left and continues_right:
        return "◀▶"
    if continues_left:
        return "◀"
    if continues_right:
        return "▶"
    return ""


def _parse_event_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return arrow.get(str(s), "YYYYMMDD").date()
    except Exception:
        return None


def _blockplan_style_rules(config: "CalendarConfig") -> list:
    """Source the style_rules list (UnifiedTheme first, legacy fallback)."""
    theme = getattr(config, "theme", None)
    if theme is not None:
        rules = theme.sections.get("style_rules")
        if isinstance(rules, list):
            return rules
    return list(getattr(config, "theme_style_rules", None) or [])


def generate_excel_blockplan(
    config: "CalendarConfig",
    db: "CalendarDB",
    out_path: Path,
) -> None:
    """Generate the Excel workbook for the ``excelblockplan`` subcommand.

    Args:
        config: Fully populated CalendarConfig.
        db:     Open CalendarDB instance.
        out_path: Destination .xlsx path (parent directory must exist).
    """
    visible_days = compute_visible_days(config)
    if not visible_days:
        return

    # freeze=False — excelblockplan rows are independent records ordered by
    # start date; sort/filter workflows want full-sheet scroll, and clearing
    # freeze_panes after the fact corrupts the XML (orphaned <selection pane>
    # elements remain), so the freeze must never be set in the first place.
    wb, ws, data_start_row, visible_days, holiday_map, right_border_cols, all_events, settings = (
        _prepare_sheet(config, db, subcommand="excelblockplan", freeze=False)
    )

    # Filter events / durations using the same predicate the other
    # visualizers use so the data sheet stays consistent with the SVG views.
    raw_dicts = db.get_all_events_in_range(
        str(config.userstart or config.adjustedstart),
        str(config.userend or config.adjustedend),
    )
    filtered_raw = filter_events(raw_dicts, config)
    # Sort the (dict, Event) pairs together so the row index used to look up
    # field values stays aligned with the sort order.
    paired: list[tuple[dict, Event]] = [
        (d, Event.from_dict(d)) for d in filtered_raw
    ]
    paired.sort(key=lambda p: (p[1].start or "", p[1].task_name or ""))
    filtered = [p[0] for p in paired]
    events = [p[1] for p in paired]

    style_engine = StyleEngine(_blockplan_style_rules(config))

    visible_start = visible_days[0]
    visible_end = visible_days[-1]
    base_font = Font(name=settings["font_name"], size=settings["font_size"])
    center = Alignment(horizontal="center", vertical="center")

    # Track which day-column cells already carry data so we can overlay (not
    # replace) holiday shading at the end.
    data_cell_argb: dict[tuple[int, int], str] = {}

    for row_offset, event in enumerate(events):
        row = data_start_row + row_offset
        ws.row_dimensions[row].height = 14

        # ── Columns A-W : events-table fields ─────────────────────────────
        for col_idx, (field_name, _w) in enumerate(FIXED_COLUMNS, start=1):
            db_key = _EVENT_FIELD_MAP.get(field_name)
            raw = filtered[row_offset].get(db_key) if db_key else None
            value = _format_cell_value(field_name, raw)
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = base_font

        fill_color, icon_color = _resolve_event_style(style_engine, event)
        ev_start = _parse_event_date(event.start)
        ev_end = _parse_event_date(event.end) or ev_start
        if ev_start is None:
            continue

        # Single-day event → icon glyph in the start-date column.
        # Multi-day duration → fill every visible day between start and end.
        if event.is_duration and ev_end and ev_end > ev_start:
            draw_start = max(ev_start, visible_start)
            draw_end = min(ev_end, visible_end)
            if draw_end < visible_start or draw_start > visible_end:
                # Entirely outside the visible range — still record the row
                # but no day-column decoration to apply.
                continues_left = ev_start < visible_start
                continues_right = ev_end > visible_end
                glyph = _continuation_glyph(
                    continues_left=continues_left,
                    continues_right=continues_right,
                )
                if glyph:
                    cont_cell = ws.cell(row=row, column=CONTINUATION_COL, value=glyph)
                    cont_cell.font = Font(
                        name=settings["font_name"],
                        size=settings["font_size"],
                        color=_font_color_argb(icon_color or fill_color),
                    )
                    cont_cell.alignment = center
                continue

            for i, d in enumerate(visible_days):
                if d < draw_start or d > draw_end:
                    continue
                col = FIRST_DATE_COL + i
                cell = ws.cell(row=row, column=col)
                cell.font = base_font
                argb = _to_argb(fill_color) if fill_color else None
                if argb is not None:
                    _apply_fill(cell, fill_color)
                    data_cell_argb[(row, col)] = argb
                if col in right_border_cols:
                    rbs = right_border_cols[col]
                    _apply_right_border(cell, rbs["style"], rbs["color"])

            continues_left = ev_start < visible_start
            continues_right = ev_end > visible_end
            glyph = _continuation_glyph(
                continues_left=continues_left,
                continues_right=continues_right,
            )
            if glyph:
                cont_cell = ws.cell(row=row, column=CONTINUATION_COL, value=glyph)
                cont_cell.font = Font(
                    name=settings["font_name"],
                    size=settings["font_size"],
                    color=_font_color_argb(icon_color or fill_color),
                )
                cont_cell.alignment = center
        else:
            # Single-day event — icon in the start column.
            col = _column_for_day(visible_days, ev_start)
            if col is None:
                continue
            glyph = _resolve_event_icon(style_engine, event)
            cell = ws.cell(row=row, column=col, value=glyph)
            cell.font = Font(
                name=settings["font_name"],
                size=settings["font_size"],
                color=_font_color_argb(icon_color or fill_color),
            )
            cell.alignment = center
            if fill_color:
                _apply_fill(cell, fill_color)
                argb = _to_argb(fill_color)
                if argb is not None:
                    data_cell_argb[(row, col)] = argb
            if col in right_border_cols:
                rbs = right_border_cols[col]
                _apply_right_border(cell, rbs["style"], rbs["color"])

    # ── Holiday / special-day overlay (drawn AFTER data rows) ─────────────
    last_row = data_start_row + len(events) - 1 if events else data_start_row - 1
    if last_row >= data_start_row:
        for row in range(data_start_row, last_row + 1):
            for i, d in enumerate(visible_days):
                if d not in holiday_map:
                    continue
                col = FIRST_DATE_COL + i
                base = data_cell_argb.get((row, col))
                cell = ws.cell(row=row, column=col)
                _apply_overlay_fill(cell, base or "", holiday_map[d]["color"])
                if col in right_border_cols:
                    rbs = right_border_cols[col]
                    _apply_right_border(cell, rbs["style"], rbs["color"])

    wb.save(str(out_path))
