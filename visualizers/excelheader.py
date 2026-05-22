"""Excel workbook generators — timeband header and blockplan-style data sheet.

This module provides the column-layout constants, helpers and timeband-row
writer that both the ``excelheader`` and ``excelblockplan`` subcommands share,
plus the ``generate_excel_header`` entry point used by the former.

Layout (shared between excelheader and excelblockplan)
------------------------------------------------------
Columns A-W : project-tracking labels — one per ``events`` table column
              ``id | status | priority | wbs | rollup | milestone |
              percent_complete | name | effort | duration | start_date |
              end_date | earliest_start_date | latest_start_date |
              earliest_end_date | latest_end_date | predecessors |
              resource_names | resource_group | notes | icon | color | tags``
Column X    : reserved for the continuation icon (filled by excelblockplan
              when a duration extends beyond the visible range)
Columns Y+  : one column per visible calendar day (width = 3 chars)
Rows 1..N   : timeband rows — heading label placed in column W,
              segment values starting at column Y
Row  N+1    : column-header row with the A-W label names
Rows N+2..  : data rows (excelheader: ``DATA_ROWS`` empty rows;
              excelblockplan: one row per event/duration)

Freeze panes are set at column Y / column-header row so the label columns
and timeband rows stay visible while scrolling.
"""

from __future__ import annotations

from bisect import bisect_left
from datetime import date, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any

import arrow
from PIL import ImageColor
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.data_models import Event
from shared.day_classifier import classify_days, day_rule_matches
from shared.icon_band import compute_icon_band_days
from visualizers.blockplan.renderer import BlockPlanRenderer, _BandSegment

if TYPE_CHECKING:
    from config.config import CalendarConfig
    from shared.db_access import CalendarDB


def _resolve_excel_token(config: "CalendarConfig", token: str) -> dict:
    """Return the unified-theme style dict for ``token`` (papersize-only ctx).

    The Excel writers have no per-event ctx (they draw timeband rows, not
    events) and no notion of paper size in the SVG sense, but ``papersize``
    is forwarded so themes can scope rules with ``select: { papersize: ... }``
    if needed. Returns ``{}`` when no theme is loaded or the token isn't
    defined.
    """
    theme = getattr(config, "theme", None)
    if theme is None:
        return {}
    ctx: dict[str, str] = {}
    papersize = getattr(config, "papersize", None)
    if papersize:
        ctx["papersize"] = str(papersize)
    return theme.resolve_token(token, ctx) or {}


# ── Constants ─────────────────────────────────────────────────────────────────

# Fixed label columns A-W — one per events-table column the user can plan around.
FIXED_COLUMNS: list[tuple[str, float]] = [
    ("id", 6),
    ("status", 9),
    ("priority", 8),
    ("wbs", 10),
    ("rollup", 7),
    ("milestone", 10),
    ("percent_complete", 10),
    ("name", 30),
    ("effort", 8),
    ("duration", 10),
    ("start_date", 12),
    ("end_date", 12),
    ("earliest_start_date", 14),
    ("latest_start_date", 14),
    ("earliest_end_date", 14),
    ("latest_end_date", 14),
    ("predecessors", 16),
    ("resource_names", 18),
    ("resource_group", 14),
    ("notes", 24),
    ("icon", 10),
    ("color", 10),
    ("tags", 12),
]
LABEL_COL_END = len(FIXED_COLUMNS)              # 23 = column W
CONTINUATION_COL = LABEL_COL_END + 1            # 24 = column X
FIRST_DATE_COL = LABEL_COL_END + 2              # 25 = column Y
DATA_ROWS = 100
DAY_COL_WIDTH = 3.0  # Excel character-width units

# ISO 3166-1 alpha-2 → flag emoji
_COUNTRY_FLAGS: dict[str, str] = {
    "ad": "🇦🇩", "ae": "🇦🇪", "af": "🇦🇫", "ag": "🇦🇬", "ai": "🇦🇮",
    "al": "🇦🇱", "am": "🇦🇲", "ao": "🇦🇴", "ar": "🇦🇷", "as": "🇦🇸",
    "at": "🇦🇹", "au": "🇦🇺", "aw": "🇦🇼", "az": "🇦🇿", "ba": "🇧🇦",
    "bb": "🇧🇧", "bd": "🇧🇩", "be": "🇧🇪", "bf": "🇧🇫", "bg": "🇧🇬",
    "bh": "🇧🇭", "bi": "🇧🇮", "bj": "🇧🇯", "bl": "🇧🇱", "bm": "🇧🇲",
    "bn": "🇧🇳", "bo": "🇧🇴", "br": "🇧🇷", "bs": "🇧🇸", "bt": "🇧🇹",
    "bw": "🇧🇼", "by": "🇧🇾", "bz": "🇧🇿", "ca": "🇨🇦", "cc": "🇨🇨",
    "cd": "🇨🇩", "cf": "🇨🇫", "cg": "🇨🇬", "ch": "🇨🇭", "ci": "🇨🇮",
    "ck": "🇨🇰", "cl": "🇨🇱", "cm": "🇨🇲", "cn": "🇨🇳", "co": "🇨🇴",
    "cr": "🇨🇷", "cu": "🇨🇺", "cv": "🇨🇻", "cy": "🇨🇾", "cz": "🇨🇿",
    "de": "🇩🇪", "dj": "🇩🇯", "dk": "🇩🇰", "dm": "🇩🇲", "do": "🇩🇴",
    "dz": "🇩🇿", "ec": "🇪🇨", "ee": "🇪🇪", "eg": "🇪🇬", "es": "🇪🇸",
    "et": "🇪🇹", "fi": "🇫🇮", "fj": "🇫🇯", "fr": "🇫🇷", "ga": "🇬🇦",
    "gb": "🇬🇧", "gd": "🇬🇩", "ge": "🇬🇪", "gh": "🇬🇭", "gm": "🇬🇲",
    "gn": "🇬🇳", "gq": "🇬🇶", "gr": "🇬🇷", "gt": "🇬🇹", "gu": "🇬🇺",
    "gw": "🇬🇼", "gy": "🇬🇾", "hk": "🇭🇰", "hn": "🇭🇳", "hr": "🇭🇷",
    "ht": "🇭🇹", "hu": "🇭🇺", "id": "🇮🇩", "ie": "🇮🇪", "il": "🇮🇱",
    "in": "🇮🇳", "iq": "🇮🇶", "ir": "🇮🇷", "is": "🇮🇸", "it": "🇮🇹",
    "jm": "🇯🇲", "jo": "🇯🇴", "jp": "🇯🇵", "ke": "🇰🇪", "kg": "🇰🇬",
    "kh": "🇰🇭", "ki": "🇰🇮", "km": "🇰🇲", "kn": "🇰🇳", "kp": "🇰🇵",
    "kr": "🇰🇷", "kw": "🇰🇼", "ky": "🇰🇾", "kz": "🇰🇿", "la": "🇱🇦",
    "lb": "🇱🇧", "lc": "🇱🇨", "li": "🇱🇮", "lk": "🇱🇰", "lr": "🇱🇷",
    "ls": "🇱🇸", "lt": "🇱🇹", "lu": "🇱🇺", "lv": "🇱🇻", "ly": "🇱🇾",
    "ma": "🇲🇦", "mc": "🇲🇨", "md": "🇲🇩", "me": "🇲🇪", "mg": "🇲🇬",
    "mh": "🇲🇭", "mk": "🇲🇰", "ml": "🇲🇱", "mm": "🇲🇲", "mn": "🇲🇳",
    "mo": "🇲🇴", "mp": "🇲🇵", "mr": "🇲🇷", "ms": "🇲🇸", "mt": "🇲🇹",
    "mu": "🇲🇺", "mv": "🇲🇻", "mw": "🇲🇼", "mx": "🇲🇽", "my": "🇲🇾",
    "mz": "🇲🇿", "na": "🇳🇦", "ne": "🇳🇪", "ng": "🇳🇬", "ni": "🇳🇮",
    "nl": "🇳🇱", "no": "🇳🇴", "np": "🇳🇵", "nr": "🇳🇷", "nu": "🇳🇺",
    "nz": "🇳🇿", "om": "🇴🇲", "pa": "🇵🇦", "pe": "🇵🇪", "pf": "🇵🇫",
    "pg": "🇵🇬", "ph": "🇵🇭", "pk": "🇵🇰", "pl": "🇵🇱", "pr": "🇵🇷",
    "ps": "🇵🇸", "pt": "🇵🇹", "pw": "🇵🇼", "py": "🇵🇾", "qa": "🇶🇦",
    "ro": "🇷🇴", "rs": "🇷🇸", "ru": "🇷🇺", "rw": "🇷🇼", "sa": "🇸🇦",
    "sb": "🇸🇧", "sc": "🇸🇨", "sd": "🇸🇩", "se": "🇸🇪", "sg": "🇸🇬",
    "sh": "🇸🇭", "si": "🇸🇮", "sk": "🇸🇰", "sl": "🇸🇱", "sm": "🇸🇲",
    "sn": "🇸🇳", "so": "🇸🇴", "sr": "🇸🇷", "ss": "🇸🇸", "st": "🇸🇹",
    "sv": "🇸🇻", "sy": "🇸🇾", "sz": "🇸🇿", "tc": "🇹🇨", "td": "🇹🇩",
    "tg": "🇹🇬", "th": "🇹🇭", "tj": "🇹🇯", "tk": "🇹🇰", "tl": "🇹🇱",
    "tm": "🇹🇲", "tn": "🇹🇳", "to": "🇹🇴", "tr": "🇹🇷", "tt": "🇹🇹",
    "tv": "🇹🇻", "tz": "🇹🇿", "ua": "🇺🇦", "ug": "🇺🇬", "us": "🇺🇸",
    "uy": "🇺🇾", "uz": "🇺🇿", "va": "🇻🇦", "vc": "🇻🇨", "ve": "🇻🇪",
    "vg": "🇻🇬", "vi": "🇻🇮", "vn": "🇻🇳", "vu": "🇻🇺", "ws": "🇼🇸",
    "ye": "🇾🇪", "za": "🇿🇦", "zm": "🇿🇲", "zw": "🇿🇼",
}

# ── Colour helpers ────────────────────────────────────────────────────────────

def _to_argb(color: str | None) -> str | None:
    """Convert any CSS color string → openpyxl ARGB hex (e.g. ``'FF4472C4'``).

    Returns ``None`` for ``None``, ``"none"``, or ``"transparent"`` so callers
    can skip applying a fill when there is nothing to render.
    """
    if not color:
        return None
    c = color.strip().lower()
    if c in {"", "none", "transparent"}:
        return None
    try:
        rgb = ImageColor.getrgb(color)
        r, g, b = rgb[0], rgb[1], rgb[2]
        return f"FF{r:02X}{g:02X}{b:02X}"
    except (ValueError, KeyError, AttributeError):
        return None


def _solid_fill(color: str | None) -> PatternFill | None:
    argb = _to_argb(color)
    if argb is None:
        return None
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _apply_fill(cell: Any, color: str | None) -> None:
    fill = _solid_fill(color)
    if fill is not None:
        cell.fill = fill


def _font_color_argb(color: str | None) -> str:
    """Return ARGB for a font color, defaulting to opaque black."""
    return _to_argb(color) or "FF000000"


# ── Segment helpers ───────────────────────────────────────────────────────────

def _group_segments(
    segments: list[_BandSegment], show_every: int
) -> list[list[_BandSegment]]:
    """Partition *segments* into consecutive groups of size *show_every*."""
    n = max(1, show_every)
    groups: list[list[_BandSegment]] = []
    buf: list[_BandSegment] = []
    for seg in segments:
        buf.append(seg)
        if len(buf) >= n:
            groups.append(buf)
            buf = []
    if buf:
        groups.append(buf)
    return groups


def _col_for_day(
    day: date, visible_days: list[date], *, end: bool = False
) -> int:
    """Return the 1-based Excel column index for *day*.

    When ``end=True`` the column corresponds to the *end_exclusive* boundary
    (i.e. the last column of a segment, inclusive).
    """
    idx = bisect_left(visible_days, day)
    if end:
        idx -= 1
    return FIRST_DATE_COL + max(0, idx)


# ── Visible-day helper ────────────────────────────────────────────────────────

def compute_visible_days(config: "CalendarConfig") -> list[date]:
    """Return the ordered list of calendar dates that get a day column.

    Honors ``config.weekend_style`` (0 = weekdays only, 1+ = full week).
    Uses ``userstart``/``userend`` if present, otherwise the adjusted range.
    """
    range_start = str(config.userstart or config.adjustedstart)
    range_end = str(config.userend or config.adjustedend)
    start = arrow.get(range_start, "YYYYMMDD").date()
    end = arrow.get(range_end, "YYYYMMDD").date()
    if end < start:
        start, end = end, start
    weekend_style = int(getattr(config, "weekend_style", 0))
    visible: list[date] = []
    cursor = start
    while cursor <= end:
        if weekend_style == 0:
            if cursor.weekday() < 5:
                visible.append(cursor)
        else:
            visible.append(cursor)
        cursor += timedelta(days=1)
    return visible


# ── Holiday pre-fetch ─────────────────────────────────────────────────────────

def _build_holiday_map(
    visible_days: list[date],
    db: "CalendarDB",
    config: "CalendarConfig",
    federal_color: str,
    company_color: str,
    weekend_color: str | None,
) -> dict[date, dict]:
    """Return a dict mapping each non-workday to display info.

    Classification uses :func:`shared.day_classifier.classify_days` so
    blockplan, excelheader and excelblockplan share one source of truth.
    Precedence when multiple classes match a single date:
    federal > company > weekend.
    """
    hmap: dict[date, dict] = {}
    classes_by_day = classify_days(visible_days, db, config)
    country = getattr(config, "country", None)
    for d in visible_days:
        classes = classes_by_day.get(d, frozenset())
        if not classes:
            continue
        daykey = d.strftime("%Y%m%d")
        if "federal_holiday" in classes:
            gov = [
                h for h in db.get_holidays_for_date(daykey, country)
                if h.get("nonworkday")
            ]
            icon_key = str((gov[0].get("icon") if gov else "") or "").lower()
            emoji = _COUNTRY_FLAGS.get(icon_key, "🏛")
            hmap[d] = {
                "color": federal_color,
                "emoji": emoji,
                "name": str((gov[0].get("displayname") if gov else "") or ""),
                "is_nonwork": True,
                "class": "federal_holiday",
            }
            continue
        if "company_holiday" in classes:
            special = [
                s for s in db.get_special_days_for_date(daykey)
                if s.get("nonworkday")
            ]
            day_color = (
                str((special[0].get("daycolor") if special else "") or "").strip()
                or company_color
            )
            icon_key = str((special[0].get("icon") if special else "") or "").lower()
            emoji = _COUNTRY_FLAGS.get(icon_key, "🏢")
            hmap[d] = {
                "color": day_color,
                "emoji": emoji,
                "name": str((special[0].get("name") if special else "") or ""),
                "is_nonwork": True,
                "class": "company_holiday",
            }
            continue
        if "weekend" in classes and weekend_color:
            hmap[d] = {
                "color": weekend_color,
                "emoji": "",
                "name": "",
                "is_nonwork": True,
                "class": "weekend",
            }
    return hmap


# ── Vertical-line → right-border mapping ─────────────────────────────────────

def _build_right_border_cols(
    vertical_lines: list[dict],
    band_segments: dict[str, list[_BandSegment]],
    visible_days: list[date],
    config: "CalendarConfig",
    *,
    default_color: str | None = None,
    default_width: float | None = None,
) -> dict[int, dict]:
    """Return {excel_col: {color, style}} for each configured vertical line."""
    tk_vline = _resolve_excel_token(config, "box:vline")
    fallback_color = default_color or "red"
    fallback_width = float(default_width if default_width is not None else 1.5)
    result: dict[int, dict] = {}
    for line in vertical_lines:
        if not isinstance(line, dict):
            continue
        band_name = str(line.get("band") or line.get("band_label") or "").strip().lower()
        value = str(line.get("value") or "").strip()
        repeat = bool(line.get("repeat", False))
        align = str(line.get("align", "end")).strip().lower()
        line_color = str(
            line.get("color")
            or tk_vline.get("stroke")
            or fallback_color
        )
        line_width = float(
            line.get("width")
            or tk_vline.get("stroke_width")
            or fallback_width
        )
        if not band_name:
            continue
        segs = band_segments.get(band_name, [])
        for seg in segs:
            if not repeat and seg.label != value:
                continue
            if align == "end":
                idx = bisect_left(visible_days, seg.end_exclusive) - 1
            elif align == "center":
                s_idx = bisect_left(visible_days, seg.start)
                e_idx = bisect_left(visible_days, seg.end_exclusive)
                idx = (s_idx + e_idx) // 2
            else:  # "start"
                idx = bisect_left(visible_days, seg.start)
            if 0 <= idx < len(visible_days):
                excel_col = FIRST_DATE_COL + idx
                result[excel_col] = {
                    "color": line_color,
                    "style": "medium" if line_width > 1.5 else "thin",
                }
    return result


def _apply_right_border(cell: Any, style: str, color: str) -> None:
    """Add (or replace) just the right border on *cell* without disturbing others."""
    argb = _to_argb(color) or "FFFF0000"
    existing = cell.border
    cell.border = Border(
        left=existing.left,
        top=existing.top,
        bottom=existing.bottom,
        right=Side(border_style=style, color=argb),
    )


def _apply_overlay_fill(cell: Any, base_argb: str, overlay_color: str | None) -> None:
    """Decorate *cell* with both an existing base fill and a holiday/special-day
    overlay color.

    When the cell already carries event/duration data (``base_argb`` non-None)
    and a non-workday colour applies (``overlay_color`` non-None), use an
    Excel pattern fill that visibly combines the two — the holiday colour
    becomes the foreground stripes, the data colour stays as the background.
    Otherwise apply a plain solid fill.  Mirrors the “show both” rule from
    the spec for blockplan-style data sheets.
    """
    overlay_argb = _to_argb(overlay_color)
    if base_argb and overlay_argb:
        cell.fill = PatternFill(
            start_color=overlay_argb,
            end_color=base_argb,
            fill_type="lightUp",
        )
        return
    if overlay_argb is not None:
        cell.fill = PatternFill(
            start_color=overlay_argb, end_color=overlay_argb, fill_type="solid"
        )


# ── Shared sheet-builder helpers ──────────────────────────────────────────────

def _read_band_settings(config: "CalendarConfig", subcommand: str) -> dict:
    """Return dict of shared excel font, colours and band defaults for *subcommand*.

    ``subcommand`` is ``"excelheader"`` or ``"excelblockplan"``.  Per-subcommand
    config-field prefixes are looked up first, falling back to ``excelheader_*``
    so themes that only set the excelheader keys also style excelblockplan
    consistently.
    """
    def _cfg(*names: str, default: Any = None) -> Any:
        for n in names:
            v = getattr(config, n, None)
            if v not in (None, ""):
                return v
        return default

    font_name = str(
        _cfg(f"{subcommand}_font_name", "excelheader_font_name", default="Calibri")
    )
    font_size = int(
        _cfg(f"{subcommand}_font_size", "excelheader_font_size", default=9)
    )
    band_row_height = float(
        _cfg(
            f"{subcommand}_band_row_height",
            "excelheader_band_row_height",
            default=18.0,
        )
    )
    header_heading_fill = str(
        _cfg(
            f"{subcommand}_header_heading_fill_color",
            "excelheader_header_heading_fill_color",
            default="none",
        )
    )
    header_label_color = str(
        _cfg(
            f"{subcommand}_header_label_color",
            "excelheader_header_label_color",
            default="black",
        )
    )
    header_label_align_h = str(
        _cfg(
            f"{subcommand}_header_label_align_h",
            "excelheader_header_label_align_h",
            default="right",
        )
    ).lower()
    timeband_fill_color = _cfg(
        f"{subcommand}_timeband_fill_color",
        "excelheader_timeband_fill_color",
        default="none",
    )
    timeband_fill_palette = _cfg(
        f"{subcommand}_timeband_fill_palette",
        "excelheader_timeband_fill_palette",
        default=[],
    ) or []
    timeband_label_color = str(
        _cfg(
            f"{subcommand}_timeband_label_color",
            "excelheader_timeband_label_color",
            default="black",
        )
    )
    federal_color = str(
        _cfg(
            f"{subcommand}_federal_holiday_fill_color",
            "excelheader_federal_holiday_fill_color",
            "theme_federal_holiday_color",
            default="#FFE4E1",
        )
    )
    company_color = str(
        _cfg(
            f"{subcommand}_company_holiday_fill_color",
            "excelheader_company_holiday_fill_color",
            "theme_company_holiday_color",
            default="#FFFACD",
        )
    )
    weekend_color = _cfg(
        f"{subcommand}_weekend_fill_color",
        "excelheader_weekend_fill_color",
        default=None,
    )
    vline_color = str(
        _cfg(
            f"{subcommand}_vertical_line_color",
            "excelheader_vertical_line_color",
            default="red",
        )
    )
    vline_width = float(
        _cfg(
            f"{subcommand}_vertical_line_width",
            "excelheader_vertical_line_width",
            default=1.5,
        )
    )
    return {
        "font_name": font_name,
        "font_size": font_size,
        "band_row_height": band_row_height,
        "header_heading_fill": header_heading_fill,
        "header_label_color": header_label_color,
        "header_label_align_h": header_label_align_h,
        "timeband_fill_color": timeband_fill_color,
        "timeband_fill_palette": timeband_fill_palette,
        "timeband_label_color": timeband_label_color,
        "federal_color": federal_color,
        "company_color": company_color,
        "weekend_color": weekend_color or None,
        "vline_color": vline_color,
        "vline_width": vline_width,
    }


def _setup_column_widths(ws: Any, visible_days: list[date]) -> None:
    """Set widths for A-W label columns, the X continuation column and Y+ days."""
    for col_idx, (_, width) in enumerate(FIXED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions[get_column_letter(CONTINUATION_COL)].width = DAY_COL_WIDTH
    for i in range(len(visible_days)):
        ws.column_dimensions[
            get_column_letter(FIRST_DATE_COL + i)
        ].width = DAY_COL_WIDTH


def _write_timebands(
    ws: Any,
    *,
    config: "CalendarConfig",
    db: "CalendarDB",
    top_bands: list[dict],
    visible_days: list[date],
    band_events: list[Event],
    holiday_map: dict[date, dict],
    day_classes: dict[date, frozenset[str]],
    band_segments: dict[str, list[_BandSegment]],
    settings: dict,
    start_row: int = 1,
) -> int:
    """Write the timeband rows and return the next free row index.

    The band heading label (``band["label"]``) is placed in column W
    (rightmost label column) by merging cells A:W and aligning per
    ``label_align_h``.  Segment cells start at column Y so the X column
    (reserved for the continuation icon) stays clear.
    """
    tk_heading = _resolve_excel_token(config, "text:heading")
    tk_band_label = _resolve_excel_token(config, "text:band_label")
    tk_box_band = _resolve_excel_token(config, "box:band")

    def _classify(d: date) -> frozenset[str]:
        return day_classes.get(d, frozenset())

    current_row = start_row
    for band in top_bands:
        band_font_name: str = str(band.get("excel_font_name") or settings["font_name"])
        band_font_size: int = int(band.get("excel_font_size") or settings["font_size"])

        row_h_pts = float(
            band.get("row_height") or settings["band_row_height"]
        )
        ws.row_dimensions[current_row].height = max(12.0, row_h_pts * 0.75)

        label_text = str(band.get("label", ""))
        heading_fill_color = str(
            band.get("label_fill_color") or settings["header_heading_fill"] or ""
        )
        heading_label_color = str(
            band.get("label_color")
            or tk_heading.get("color")
            or settings["header_label_color"]
        )
        heading_align_h = str(
            band.get("label_align_h") or settings["header_label_align_h"]
        ).lower()
        excel_h_align = (
            "right" if heading_align_h == "right"
            else "center" if heading_align_h == "center"
            else "left"
        )

        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=LABEL_COL_END,
        )
        heading_cell = ws.cell(row=current_row, column=1, value=label_text)
        heading_cell.font = Font(
            name=band_font_name,
            size=band_font_size,
            bold=True,
            color=_font_color_argb(heading_label_color),
        )
        heading_cell.alignment = Alignment(
            horizontal=excel_h_align, vertical="center", wrap_text=False
        )
        _apply_fill(heading_cell, heading_fill_color)

        # ── Icon band — one cell per visible day in Y+ ────────────────────
        if str(band.get("unit", "")).strip().lower() == "icon":
            icon_rules = list(band.get("icon_rules") or [])
            day_icon_map = compute_icon_band_days(
                band_events, icon_rules, visible_days, classify_fn=_classify
            )
            icon_fill = str(band.get("fill_color") or "none")
            for i, d in enumerate(visible_days):
                col = FIRST_DATE_COL + i
                icons = day_icon_map.get(d, [])
                if d in holiday_map:
                    _apply_fill(ws.cell(row=current_row, column=col), holiday_map[d]["color"])
                elif icon_fill and icon_fill.lower() not in {"none", "transparent"}:
                    _apply_fill(ws.cell(row=current_row, column=col), icon_fill)
                if not icons:
                    continue
                icon_name, icon_color = icons[0]
                symbol = "●"  # ● filled circle
                cell = ws.cell(row=current_row, column=col, value=symbol)
                cell.font = Font(
                    name=band_font_name,
                    size=band_font_size,
                    color=_font_color_argb(icon_color),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            continue

        segs = band_segments.get(str(band.get("label", "")).strip().lower(), [])

        band_fill_raw = band.get(
            "fill_color",
            tk_box_band.get("fill") or settings["timeband_fill_color"],
        )
        band_palette_raw = band.get("fill_palette", settings["timeband_fill_palette"])
        color_list = BlockPlanRenderer._resolve_color_list(
            band_fill_raw, band_palette_raw, db
        )

        seg_label_color = str(
            band.get("font_color")
            or tk_band_label.get("color")
            or settings["timeband_label_color"]
        )
        show_every = max(1, int(band.get("show_every", 1)))
        label_values: list | None = band.get("label_values")
        band_fill_rules_raw = band.get("fill_rules")
        band_fill_rules: list[dict] | None = (
            band_fill_rules_raw if isinstance(band_fill_rules_raw, list) else None
        )
        groups = _group_segments(segs, show_every)

        for gidx, group in enumerate(groups):
            seg_start = group[0].start
            seg_end_excl = group[-1].end_exclusive

            col_s = _col_for_day(seg_start, visible_days)
            col_e = _col_for_day(seg_end_excl, visible_days, end=True)

            col_s = max(FIRST_DATE_COL, min(col_s, FIRST_DATE_COL + len(visible_days) - 1))
            col_e = max(col_s, min(col_e, FIRST_DATE_COL + len(visible_days) - 1))

            if label_values and gidx < len(label_values):
                cell_text: str = str(label_values[gidx] or group[0].label)
            else:
                cell_text = group[0].label

            is_single_day = (col_s == col_e)
            cell_fill_color: str | None = (
                color_list[gidx % len(color_list)] if color_list else None
            )
            if is_single_day:
                day_idx = col_s - FIRST_DATE_COL
                if 0 <= day_idx < len(visible_days):
                    d = visible_days[day_idx]
                    if band_fill_rules:
                        matched = False
                        for rule in band_fill_rules:
                            if not isinstance(rule, dict):
                                continue
                            match = rule.get("match") or {}
                            if isinstance(match, dict) and day_rule_matches(
                                _classify(d), match
                            ):
                                color = rule.get("color")
                                if color:
                                    cell_fill_color = str(color)
                                matched = True
                                break
                        if not matched and d in holiday_map:
                            cell_fill_color = holiday_map[d]["color"]
                            if holiday_map[d]["emoji"]:
                                cell_text = holiday_map[d]["emoji"]
                    elif d in holiday_map:
                        cell_fill_color = holiday_map[d]["color"]
                        if holiday_map[d]["emoji"]:
                            cell_text = holiday_map[d]["emoji"]

            if col_e > col_s:
                ws.merge_cells(
                    start_row=current_row, start_column=col_s,
                    end_row=current_row, end_column=col_e,
                )
            seg_cell = ws.cell(row=current_row, column=col_s, value=cell_text)
            seg_cell.font = Font(
                name=band_font_name,
                size=band_font_size,
                color=_font_color_argb(seg_label_color),
            )
            seg_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            _apply_fill(seg_cell, cell_fill_color)

        current_row += 1

    return current_row


def _write_column_header_row(
    ws: Any,
    *,
    header_row: int,
    config: "CalendarConfig",
    visible_days: list[date],
    holiday_map: dict[date, dict],
    right_border_cols: dict[int, dict],
    settings: dict,
) -> None:
    """Write the A-W label row, then apply holiday shading / vertical-line borders."""
    header_font = Font(
        name=settings["font_name"], size=settings["font_size"], bold=True
    )
    header_align_center = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[header_row].height = 18

    for col_idx, (label, _) in enumerate(FIXED_COLUMNS, start=1):
        hcell = ws.cell(row=header_row, column=col_idx, value=label)
        hcell.font = header_font
        hcell.alignment = header_align_center

    # X column header — blank but anchored so column dimensions track.
    _x_cell = ws.cell(row=header_row, column=CONTINUATION_COL, value="")
    _x_cell.alignment = header_align_center

    for i, d in enumerate(visible_days):
        col = FIRST_DATE_COL + i
        hcell = ws.cell(row=header_row, column=col)
        hcell.font = Font(name=settings["font_name"], size=settings["font_size"])
        hcell.alignment = header_align_center
        if d in holiday_map:
            _apply_fill(hcell, holiday_map[d]["color"])
        if col in right_border_cols:
            rbs = right_border_cols[col]
            _apply_right_border(hcell, rbs["style"], rbs["color"])


def _prepare_sheet(
    config: "CalendarConfig",
    db: "CalendarDB",
    *,
    subcommand: str,
) -> tuple[Any, Any, int, list[date], dict[date, dict], dict[int, dict], list[Event], dict]:
    """Build a workbook, write timeband + column-header rows, return shared state.

    Returns (workbook, worksheet, data_start_row, visible_days, holiday_map,
    right_border_cols, all_events_objects, settings).
    The data_start_row is the first row available for callers to write data
    (one row past the column-header row).  ``all_events_objects`` are the
    Event dataclasses sourced for icon-band evaluation; callers can reuse
    them for downstream rendering.
    """
    visible_days = compute_visible_days(config)
    settings = _read_band_settings(config, subcommand)

    top_bands_field = f"{subcommand}_top_time_bands"
    vlines_field = f"{subcommand}_vertical_lines"
    top_bands: list[dict] = list(
        getattr(config, top_bands_field, None)
        or getattr(config, "excelheader_top_time_bands", [])
        or []
    )
    vertical_lines: list[dict] = list(
        getattr(config, vlines_field, None)
        or getattr(config, "excelheader_vertical_lines", [])
        or []
    )

    holiday_map = _build_holiday_map(
        visible_days, db, config,
        settings["federal_color"], settings["company_color"], settings["weekend_color"],
    )
    day_classes = classify_days(visible_days, db, config)

    # Always source events — icon bands need them and excelblockplan needs
    # them for data rows.  When no events exist this is a cheap call.
    range_start_str = str(config.userstart or config.adjustedstart)
    range_end_str = str(config.userend or config.adjustedend)
    raw_events = db.get_all_events_in_range(range_start_str, range_end_str)
    band_events: list[Event] = [
        Event.from_dict(e) if isinstance(e, dict) else e for e in raw_events
    ]

    # Segments cached for both heading rendering and vertical-line lookup.
    _renderer = BlockPlanRenderer()
    range_start = str(config.userstart or config.adjustedstart)
    range_end = str(config.userend or config.adjustedend)
    start = arrow.get(range_start, "YYYYMMDD").date()
    end = arrow.get(range_end, "YYYYMMDD").date()
    if end < start:
        start, end = end, start
    band_segments: dict[str, list[_BandSegment]] = {}
    for band in top_bands:
        bname = str(band.get("label", "")).strip().lower()
        if bname and str(band.get("unit", "")).strip().lower() != "icon":
            band_segments[bname] = _renderer._build_segments(
                band, start, end, config, visible_days=visible_days, db=db
            )

    right_border_cols = _build_right_border_cols(
        vertical_lines, band_segments, visible_days, config,
        default_color=settings["vline_color"],
        default_width=settings["vline_width"],
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planner"

    _setup_column_widths(ws, visible_days)

    header_row = _write_timebands(
        ws,
        config=config, db=db,
        top_bands=top_bands,
        visible_days=visible_days,
        band_events=band_events,
        holiday_map=holiday_map,
        day_classes=day_classes,
        band_segments=band_segments,
        settings=settings,
        start_row=1,
    )

    _write_column_header_row(
        ws,
        header_row=header_row,
        config=config,
        visible_days=visible_days,
        holiday_map=holiday_map,
        right_border_cols=right_border_cols,
        settings=settings,
    )

    data_start_row = header_row + 1

    # Freeze pane: keep label cols + timeband rows visible.
    ws.freeze_panes = f"{get_column_letter(FIRST_DATE_COL)}{header_row}"

    return wb, ws, data_start_row, visible_days, holiday_map, right_border_cols, band_events, settings


# ── Main entry point — excelheader ───────────────────────────────────────────

def generate_excel_header(
    config: "CalendarConfig",
    db: "CalendarDB",
    out_path: Path,
) -> None:
    """Generate the Excel workbook for the ``excelheader`` subcommand.

    Produces the shared A-W / Y+ skeleton plus ``DATA_ROWS`` empty data rows
    decorated with holiday shading and vertical-line borders.

    Args:
        config: Fully populated CalendarConfig (date range + theme applied).
        db:     Open CalendarDB instance.
        out_path: Destination .xlsx path (parent directory must exist).
    """
    visible_days = compute_visible_days(config)
    if not visible_days:
        return

    wb, ws, data_start, visible_days, holiday_map, right_border_cols, _events, _settings = (
        _prepare_sheet(config, db, subcommand="excelheader")
    )

    for row in range(data_start, data_start + DATA_ROWS):
        ws.row_dimensions[row].height = 14
        ws.cell(row=row, column=1).value = ""
        for i, d in enumerate(visible_days):
            col = FIRST_DATE_COL + i
            if d in holiday_map:
                dcell = ws.cell(row=row, column=col)
                _apply_fill(dcell, holiday_map[d]["color"])
        for col, rbs in right_border_cols.items():
            dcell = ws.cell(row=row, column=col)
            _apply_right_border(dcell, rbs["style"], rbs["color"])

    wb.save(str(out_path))
